"""Genera Excel i PDF amb el resum d'hores del projecte."""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# Dades — hores desenvolupament + proves + altres
# ============================================================
# (data, inici, fi, commits, h_dev, h_proves, h_altres, bloc)
JORNADES = [
    ("30/03/2026", "10:19", "",      1,  2.0, 1.5, 0.0, "Versio inicial del motor"),
    ("31/03/2026", "15:27", "",      1,  2.0, 2.0, 0.0, "Implementacio V4: regles RF1-RF9"),
    ("01/04/2026", "08:53", "16:42", 6,  6.5, 1.5, 0.0, "UI, cerca comandes, tipus descarrega, llista progressiva"),
    ("02/04/2026", "08:32", "",      1,  2.0, 1.5, 0.0, "UX, optimitzacio ompliment palets, agrupacio comandes"),
    ("03/04/2026", "",      "",      0,  0.0, 5.0, 3.0, "Proves funcionals regles, analisi model de dades SQL"),
    ("04/04/2026", "",      "",      0,  0.0, 4.0, 4.0, "Proves amb comandes reals, planificacio i documentacio"),
    ("07/04/2026", "10:16", "17:06", 14, 6.0, 2.0, 0.0, "RF4/RF6, filtres, polling, preview palets, mode fosc"),
    ("08/04/2026", "11:35", "17:03", 7,  5.5, 2.5, 0.0, "Migracio CPALBARA/ALBLINIA, desplegament servidor"),
    ("09/04/2026", "07:55", "14:42", 10, 5.5, 2.5, 0.0, "Fixes palets, filtres serie, total unitats, cerca"),
    ("10/04/2026", "",      "",      0,  0.0, 5.5, 2.5, "Proves validacio migracio, comunicacio i seguiment"),
    ("14/04/2026", "11:47", "",      1,  2.0, 1.0, 0.0, "Fix divisio article en palet mixt cross-base"),
    ("22/04/2026", "15:09", "17:04", 10, 3.0, 2.5, 0.0, "Regles V6, pedido original, max individual per article"),
    ("23/04/2026", "08:06", "14:33", 12, 5.5, 2.5, 0.0, "Aprovisionament estoc, seguretat admin, lookup KAIS"),
    ("24/04/2026", "",      "",      0,  0.0, 5.0, 3.0, "Proves regles V6 amb casos reals, reunions validacio"),
    ("27/04/2026", "10:14", "15:13", 9,  5.0, 3.0, 0.0, "Optimitzacio SQL, monitoritzacio, cache, single-flight"),
    ("28/04/2026", "08:26", "17:08", 10, 6.0, 2.0, 0.0, "Connexions BD, pool, semaphore, background refresh"),
    ("29/04/2026", "07:58", "16:33", 33, 5.5, 2.5, 0.0, "Regles V7, 4 rondes optimitzacio SQL, batch endpoint, fixes"),
]

TOTAL_DIES = len(JORNADES)
TOTAL_COMMITS = sum(j[3] for j in JORNADES)
TOTAL_H_DEV = sum(j[4] for j in JORNADES)
TOTAL_H_PROVES = sum(j[5] for j in JORNADES)
TOTAL_H_ALTRES = sum(j[6] for j in JORNADES)
TOTAL_HORES = TOTAL_H_DEV + TOTAL_H_PROVES + TOTAL_H_ALTRES

AREES = [
    ("Motor de calcul i regles de negoci", 30,
     "Regles RF1-RF9 (filtratge, minim, aprovisionament, especials, general)\n"
     "Regles V6: pedido original, max individual, resolucio client\n"
     "Regles V7: dimensio especial, sac colagne normal, RF8\n"
     "Optimitzacio ompliment palets, cross-base, capes parcials\n"
     "Agrupacio de comandes (multiples comandes = 1 calcul)"),
    ("Interficie d'usuari", 18,
     "Llista comandes pendents amb calcul automatitzat d'estats\n"
     "Filtres multi-select (serie, magatzem, client)\n"
     "Preview visual, mode fosc, ajuda, exportacio CSV\n"
     "Cache localStorage, polling intel.ligent amb fingerprint"),
    ("Integracio BD i rendiment", 26,
     "Migracio ek_Pedido a CPALBARA/ALBLINIA\n"
     "4 rondes optimitzacio SQL: caches TTL/LRU, batch queries\n"
     "Endpoint /api/calcular-batch, SET NOCOUNT ON, CTE\n"
     "Single-flight locks, monitoritzacio SQL amb dashboard"),
    ("Infraestructura i desplegament", 8,
     "Desplegament servidor, seguretat endpoints admin\n"
     "Control concurrencia: semaphore 1 connexio\n"
     "Tracking activitat usuaris"),
    ("Proves i validacio", 22,
     "Proves funcionals amb comandes reals\n"
     "Validacio dades mestres amb avisos\n"
     "Trazabilitat completa de cada decisio\n"
     "Coherencia caches, cicles fix-prova"),
    ("Analisi, reunions i documentacio", 12,
     "Analisi previa: taules SQL, model de dades, planificacio\n"
     "Comunicacio: correus, discussio requisits, validacio casos\n"
     "Documentacio: guies desplegament, pagina ajuda"),
]

# ============================================================
# Excel
# ============================================================
def generar_excel():
    wb = openpyxl.Workbook()

    # Estils
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(bold=True, size=14, color="2F5496")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    subtotal_fill = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def style_cell(cell, align="left"):
        cell.border = thin_border
        cell.alignment = Alignment(horizontal=align, vertical="center")

    # --- Fulla 1: Detall per jornada ---
    ws1 = wb.active
    ws1.title = "Detall Jornades"

    ws1.merge_cells("A1:H1")
    ws1["A1"] = "Motor de Preparacio de Comandes de Venda - Resum Desenvolupament"
    ws1["A1"].font = title_font

    ws1.merge_cells("A2:H2")
    ws1["A2"] = (f"Periode: 30/03/2026 - 29/04/2026  |  {TOTAL_DIES} dies  |  "
                 f"{TOTAL_HORES:.0f} hores totals  |  {TOTAL_COMMITS} commits")
    ws1["A2"].font = Font(size=10, italic=True)

    headers = ["Data", "Inici", "Fi", "Commits", "H. Desenv.", "H. Proves", "H. Altres", "H. Total", "Bloc de treball"]
    row = 4
    for c, h in enumerate(headers, 1):
        ws1.cell(row=row, column=c, value=h)
    style_header(ws1, row, len(headers))

    for j in JORNADES:
        row += 1
        h_total = j[4] + j[5] + j[6]
        vals = [j[0], j[1] if j[1] else "-", j[2] if j[2] else "-", j[3], j[4], j[5], j[6], h_total, j[7]]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=c, value=v)
            align = "center" if c <= 4 else ("right" if c <= 8 else "left")
            style_cell(cell, align)

    # Fila subtotal jornades
    row += 1
    ws1.cell(row=row, column=1, value="Subtotal jornades")
    ws1.cell(row=row, column=4, value=TOTAL_COMMITS)
    ws1.cell(row=row, column=5, value=TOTAL_H_DEV)
    ws1.cell(row=row, column=6, value=TOTAL_H_PROVES)
    ws1.cell(row=row, column=7, value=TOTAL_H_ALTRES)
    ws1.cell(row=row, column=8, value=TOTAL_HORES)
    for c in range(1, 10):
        cell = ws1.cell(row=row, column=c)
        cell.font = total_font
        cell.fill = total_fill
        style_cell(cell, "center" if c <= 4 else "right")

    # Amplades
    for col, w in [("A", 14), ("B", 8), ("C", 8), ("D", 10),
                   ("E", 12), ("F", 12), ("G", 10), ("H", 10), ("I", 60)]:
        ws1.column_dimensions[col].width = w

    # --- Fulla 2: Resum per area ---
    ws2 = wb.create_sheet("Resum per Area")

    ws2.merge_cells("A1:C1")
    ws2["A1"] = "Distribucio d'hores per area funcional"
    ws2["A1"].font = title_font

    headers2 = ["Area", "Hores", "Detall"]
    row = 3
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=row, column=c, value=h)
    style_header(ws2, row, len(headers2))

    for area, hores, detall in AREES:
        row += 1
        for c, v in enumerate([area, hores, detall], 1):
            cell = ws2.cell(row=row, column=c, value=v)
            align = "right" if c == 2 else "left"
            style_cell(cell, align)
            cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=(c == 3))

    row += 1
    ws2.cell(row=row, column=1, value="TOTAL")
    ws2.cell(row=row, column=2, value=sum(a[1] for a in AREES))
    for c in range(1, 4):
        cell = ws2.cell(row=row, column=c)
        cell.font = total_font
        cell.fill = total_fill
        style_cell(cell, "center" if c == 1 else "right")

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 60

    for r in range(4, row):
        ws2.row_dimensions[r].height = 60

    path = os.path.join(OUTPUT_DIR, "Resum_Hores_Projecte.xlsx")
    wb.save(path)
    print(f"Excel generat: {path}")
    return path


# ============================================================
# PDF
# ============================================================
class InformePDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 9)
        self.set_text_color(100, 100, 100)
        self.cell(0, 5, "Motor de Preparacio de Comandes de Venda", align="L")
        self.cell(0, 5, "Resum de Desenvolupament", align="R", new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(47, 84, 150)
        self.set_line_width(0.5)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(4)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(150, 150, 150)
        self.cell(0, 10, f"Pagina {self.page_no()}/{{nb}}", align="C")


def generar_pdf():
    pdf = InformePDF()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # Titol
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(47, 84, 150)
    pdf.cell(0, 12, "Motor de Preparacio de Comandes de Venda", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 8, "Informe de Desenvolupament", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    # Resum executiu
    pdf.set_fill_color(214, 228, 240)
    pdf.set_draw_color(47, 84, 150)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "  Resum Executiu", fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.ln(2)

    resum = [
        ("Periode", "30 de marc - 29 d'abril de 2026"),
        ("Dies treballats", str(TOTAL_DIES)),
        ("Hores desenvolupament", f"{TOTAL_H_DEV:.0f}"),
        ("Hores proves i validacio", f"{TOTAL_H_PROVES:.0f}"),
        ("Hores analisi/reunions/docs", f"{TOTAL_H_ALTRES:.0f}"),
        ("Total hores projecte", f"{TOTAL_HORES:.0f}"),
        ("Total commits", str(TOTAL_COMMITS)),
        ("Tecnologies", "Python 3, Flask, SQL Server, JavaScript"),
    ]
    for label, val in resum:
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(55, 6, f"  {label}:", new_x="END")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, val, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    # Taula detall jornades
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(47, 84, 150)
    pdf.cell(0, 8, "Detall per Jornada", new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)

    col_w = [22, 16, 11, 11, 11, 11, 11, 77]
    headers = ["Data", "Horari", "Commits", "H.Dev", "H.Test", "H.Altre", "H.Tot", "Bloc de treball"]
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(47, 84, 150)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)
    fill = False
    for j in JORNADES:
        horari = f"{j[1]}-{j[2]}" if j[1] and j[2] else (j[1] if j[1] else "-")
        h_total = j[4] + j[5] + j[6]
        vals = [j[0], horari, str(j[3]), f"{j[4]:.1f}", f"{j[5]:.1f}", f"{j[6]:.1f}", f"{h_total:.1f}", j[7]]
        aligns = ["C", "C", "C", "C", "C", "C", "C", "L"]
        if fill:
            pdf.set_fill_color(240, 244, 249)
        for i, v in enumerate(vals):
            pdf.cell(col_w[i], 6, v, border=1, fill=fill, align=aligns[i])
        pdf.ln()
        fill = not fill

    # Fila total
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(214, 228, 240)
    pdf.cell(col_w[0] + col_w[1], 7, "TOTAL", border=1, fill=True, align="C")
    pdf.cell(col_w[2], 7, str(TOTAL_COMMITS), border=1, fill=True, align="C")
    pdf.cell(col_w[3], 7, f"{TOTAL_H_DEV:.0f}", border=1, fill=True, align="C")
    pdf.cell(col_w[4], 7, f"{TOTAL_H_PROVES:.0f}", border=1, fill=True, align="C")
    pdf.cell(col_w[5], 7, f"{TOTAL_H_ALTRES:.0f}", border=1, fill=True, align="C")
    pdf.cell(col_w[6], 7, f"{TOTAL_HORES:.0f}", border=1, fill=True, align="C")
    pdf.cell(col_w[7], 7, "", border=1, fill=True)
    pdf.ln(10)

    # Distribucio per area
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(47, 84, 150)
    pdf.cell(0, 8, "Distribucio per Area Funcional", new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)

    col_w2 = [55, 15, 100]
    headers2 = ["Area", "Hores", "Detall"]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(47, 84, 150)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(headers2):
        pdf.cell(col_w2[i], 7, h, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8.5)
    for area, hores, detall in AREES:
        detall_lines = detall.replace("\n", " | ")
        line_h = 6
        pdf.cell(col_w2[0], line_h, area, border="LTB", align="L")
        pdf.cell(col_w2[1], line_h, str(hores), border="TB", align="C")
        pdf.set_font("Helvetica", "", 7.5)
        pdf.cell(col_w2[2], line_h, detall_lines[:120], border="RTB", align="L")
        pdf.set_font("Helvetica", "", 8.5)
        pdf.ln()

    # Total
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(214, 228, 240)
    pdf.cell(col_w2[0], 7, "TOTAL", border=1, fill=True, align="C")
    pdf.cell(col_w2[1], 7, str(sum(a[1] for a in AREES)), border=1, fill=True, align="C")
    pdf.cell(col_w2[2], 7, "", border=1, fill=True)
    pdf.ln(10)

    # Metriques rendiment
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(47, 84, 150)
    pdf.cell(0, 8, "Metriques de Rendiment Actuals", new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 9)

    metriques = [
        ("Temps mitja per query SQL", "3 ms"),
        ("Queries amb cache", "100% (TTL 10 min o permanent)"),
        ("Connexions simultanees max", "1"),
        ("Polling interval", "60-75 s (amb jitter)"),
        ("Errors SQL", "0"),
    ]
    for label, val in metriques:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(60, 6, f"  {label}:", new_x="END")
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 6, val, new_x="LMARGIN", new_y="NEXT")

    path = os.path.join(OUTPUT_DIR, "Resum_Hores_Projecte.pdf")
    pdf.output(path)
    print(f"PDF generat: {path}")
    return path


if __name__ == "__main__":
    generar_excel()
    generar_pdf()
